# coding=utf-8

import urllib
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import requests
from selenium.webdriver.support.ui import Select


def get_all_num(url):
    file = r"F:/cpe/cpeState.xlsx"
    wb = openpyxl.load_workbook(file)
    sheet = wb["Sheet1"]
    sleep(5)
    s1 = Select(driver.find_element_by_xpath(
        '//*[@id="page-content"]/div/div/div/div[2]/div/div/div[1]/enterprise-ce/div[2]/table/tfoot/tr[1]/td/div/nav/div[1]/div/select'))
    i = 0
    for j in range(2, 47):
        sleep(10)
        html = driver.execute_script("return document.documentElement.outerHTML")
        print(html)
        reg1 = r'<span uib-tooltip="已离线" ng-if="row.onlineStatus == .*" class="badge badge-danger">离</span>'  # 返回为一列表
        imgre = re.compile(reg1)
        nums = re.findall(imgre, html)
        print(nums)


        for num in nums:
            # reg2 = r'[a-z0-9]{12}'
            searchObj = re.search(r'[a-z0-9]{12}', num)
            print(searchObj)
            sheet.cell(i + 2, 2).value = searchObj.group()
            i = i + 1
            s1.select_by_index(j)
            # bg.append(searchObj.group())
            # print(bg)
            # sheet.cell(i+2, 1).value = num
            # reg2 = r'[\u4e00-\u9fa5]{3}'
            # status = re.match(r'[\u4e00-\u9fa5]{3}',html)
            # if matchObj:
            # sheet.cell(i+2, 2).value = status
    wb.save(file)
    wb.close()



if __name__ == '__main__':
    driver = webdriver.Chrome()
    ele = driver.get('http://221.6.205.118:9080/#/device/cpe')
    driver.find_element_by_id("inputEmail").send_keys("CmpUser")
    driver.find_element_by_id("inputPassword").send_keys("Root@123")
    sleep(10)
    driver.find_element_by_xpath("//form/div/button").click()
    url = "http://221.6.205.118:9080/#/device/cpe'"
    get_all_num(url)